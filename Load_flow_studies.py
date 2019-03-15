# Copyright © 2018 Mads-Emil Kvammen <mekvammen@gmail.com>
#
# Distributed under terms of the MIT license.
# https://no.wikipedia.org/wiki/MIT-lisensen

import openpyxl
from Load_flow_calculations import Load_flow
def Print_all(BusInfo, total, S, loss, total_loss):

    print('\nBus information \n-----------------')
    print('      |           Voltage             |        Generator data         |          Load data          |')
    print('#Bus  |     V [pu]    |  Angle [deg]  |     P [pu]    |     Q [pu]    |    P [pu]    |    Q [pu]    |')

    for k in range(len(BusInfo)):
        print('{}     |  {}  |  {}  |  {}  |  {}  |  {}  |  {}  |'.format(k + 1, format(BusInfo[k,0], '+.8f'),
                                                        format(BusInfo[k,1], '+.8f'),
                                                        format(BusInfo[k,2], '+.8f'),
                                                        format(BusInfo[k,3], '+.8f'),
                                                        format(abs(BusInfo[k,4]), '+.7f'),
                                                        format(abs(BusInfo[k,5]), '+.7f')))

    print('----------------------------------------------------------------------------------------------------')
    print('                          Total:      |  {}  |  {}  |  {}  |  {}  |'.format(format(total[0], '+.8f'),
                                                  format(total[1], '+.8f'),
                                                  format(abs(total[2]), '+.7f'),
                                                  format(abs(total[3]), '+.7f')))

    print('\nBranch information \n-------------------')
    print('From | To  |       From Bus Injection      |        To Bus Injection      |          Line loss          |')
    print('bus  | bus |     P [pu]    |     Q [pu]    |     P [pu]    |     Q [pu]   |    P [pu]    |    Q [pu]    |')
    for k in range(0,len(S)):
        print('  {}  |  {}  |  {:+.8f}  |  {:+0.8f}  |  {:+0.8f}  |  {:+0.8f} |  {:+0.8f} |  {:+0.8f} |'.format(int(S[k,0]),
                                                     int(S[k,1]),S[k,2],S[k,3],S[k,4],S[k,5],loss[k, 2],loss[k, 3]))
    print('--------------------------------------------------------------------------------------------------------')
    print('                                                              Total:      |  {} |  {} |'.format(format(total_loss[0], '+.8f'),
                                                                                       format(total_loss[1], '+.8f')))
    return None

def save_to_excel(BusInfo, S, loss, YbusG, YbusB, total, total_loss):
    # Create workbook object
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = 'Bus_information'
    # Add titles in the first row of each column
    sheet.cell(row=2, column=1).value = 'Bus'
    sheet.cell(row=1, column=2).value = 'Voltage'
    sheet.cell(row=2, column=2).value = 'V [pu]'
    sheet.cell(row=2, column=3).value = 'Angle [deg]'
    sheet.cell(row=1, column=4).value = 'Generator data'
    sheet.cell(row=2, column=4).value = 'P [pu]'
    sheet.cell(row=2, column=5).value = 'Q [pu]'
    sheet.cell(row=1, column=6).value = 'Load data'
    sheet.cell(row=2, column=6).value = 'P [pu]'
    sheet.cell(row=2, column=7).value = 'Q [pu]'

    # Loop to set the value of each cell
    for k in range(0, len(BusInfo)):
        sheet.cell(row=k + 3, column=1).value = k+1
        sheet.cell(row=k + 3, column=2).value = BusInfo[k,0]
        sheet.cell(row=k + 3, column=3).value = BusInfo[k,1]
        sheet.cell(row=k + 3, column=4).value = BusInfo[k,2]
        sheet.cell(row=k + 3, column=5).value = BusInfo[k,3]
        sheet.cell(row=k + 3, column=6).value = abs(BusInfo[k,4])
        sheet.cell(row=k + 3, column=7).value = abs(BusInfo[k,5])

    sheet.cell(row=k + 5, column=3).value = 'Total:'
    for j in range(0,len(total)):
        sheet.cell(row=k + 5, column=j+4).value = abs(total[j])

    # Create new Excel-sheet
    sheet2 = wb.create_sheet()
    sheet2.title = 'Branch_information'
    # Add titles in the first row of each column
    sheet2.cell(row=1, column=1).value = 'From'
    sheet2.cell(row=2, column=1).value = 'bus'
    sheet2.cell(row=1, column=2).value = 'To'
    sheet2.cell(row=2, column=2).value = 'bus'
    sheet2.cell(row=1, column=3).value = 'From Bus Injection'
    sheet2.cell(row=2, column=3).value = 'P [pu]'
    sheet2.cell(row=2, column=4).value = 'Q [pu]'
    sheet2.cell(row=1, column=5).value = 'To Bus Injection'
    sheet2.cell(row=2, column=5).value = 'P [pu]'
    sheet2.cell(row=2, column=6).value = 'Q [pu]'
    sheet2.cell(row=1, column=7).value = 'Line loss'
    sheet2.cell(row=2, column=7).value = 'P [pu]'
    sheet2.cell(row=2, column=8).value = 'Q [pu]'


    # Loop to set the value of each cell
    for k in range(0, len(S)):
        sheet2.cell(row=k + 3, column=1).value = int(S[k,0])
        sheet2.cell(row=k + 3, column=2).value = int(S[k,1])
        sheet2.cell(row=k + 3, column=3).value = S[k, 2]
        sheet2.cell(row=k + 3, column=4).value = S[k, 3]
        sheet2.cell(row=k + 3, column=5).value = S[k, 4]
        sheet2.cell(row=k + 3, column=6).value = S[k, 5]
        sheet2.cell(row=k + 3, column=7).value = loss[k, 2]
        sheet2.cell(row=k + 3, column=8).value = loss[k, 3]
    sheet2.cell(row=k + 5, column=6).value = 'Total:'
    for j in range(0, len(total_loss)):
        sheet2.cell(row=k + 5, column=j + 7).value = total_loss[j]

    # Create new Excel-sheet
    sheet3 = wb.create_sheet()
    sheet3.title = 'Ybus_real_part'
    for i in range(0,len(YbusG)):
        for j in range(0,len(YbusG)):
            sheet3.cell(row=i + 1, column=j+1).value = YbusG[i, j]

    # Create new Excel-sheet
    sheet4 = wb.create_sheet()
    sheet4.title = 'Ybus_imaginary_part'
    for i in range(0, len(YbusB)):
        for j in range(0, len(YbusB)):
            sheet4.cell(row=i + 1, column=j + 1).value = YbusB[i, j]

    wb.save('LoadFlow_results.xlsx')
    return None

def main():

    ## Initialisation
    filenam = 'OurData.xlsx'
    branch_data_sheet = 'BranchData'
    bus_data_sheet = 'BusData'

    BusInfo, total, S, loss, total_loss, Ybus, YbusG, YbusB = Load_flow(filenam, branch_data_sheet, bus_data_sheet)

    #print('\nY-bus matrix: \n{}'.format(Ybus))
    Print_all(BusInfo, total, S, loss, total_loss)

    save_to_excel(BusInfo, S, loss, YbusG, YbusB, total, total_loss)


if __name__ == '__main__':
    main()